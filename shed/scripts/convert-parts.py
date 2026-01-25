#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""Convert shed_cut_sheet_with_ids.xlsx to parts-data.js"""

import json
from pathlib import Path

from openpyxl import load_workbook

def main():
    script_dir = Path(__file__).parent
    data_file = script_dir.parent / "data" / "shed_cut_sheet_with_ids.xlsx"
    output_file = script_dir.parent / "js" / "parts-data.js"

    wb = load_workbook(data_file)
    ws = wb.active

    # Get headers from first row
    headers = [cell.value for cell in ws[1]]

    # Valid piece ID prefixes (actual parts, not summary rows)
    valid_prefixes = (
        "F-BLK-", "FL-", "BW-", "LW-", "RW-", "FW-",
        "RF-", "SH-", "SD-", "TR-", "DR-"
    )

    parts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip section header rows (no Piece ID)
        if row[0] is None:
            continue

        # Skip summary rows - only include valid piece IDs
        piece_id = str(row[0])
        if not piece_id.startswith(valid_prefixes):
            continue

        part = {
            "id": piece_id,
            "description": row[1] or "",
            "material": row[2] or "",
            "length": str(row[3]) if row[3] else "",
            "qty": row[4] if row[4] else 1,
            "step": row[5] if row[5] else 0,
            "notes": row[6] or ""
        }
        parts.append(part)

    # Write as JS module
    js_content = f"// Generated from shed_cut_sheet_with_ids.xlsx - DO NOT EDIT\nconst partsData = {json.dumps(parts, indent=2)};\n"

    output_file.write_text(js_content)
    print(f"Generated {output_file} with {len(parts)} parts")

if __name__ == "__main__":
    main()
